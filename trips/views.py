"""
Views for Trips application with permission checks
"""
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Sum, F, Case, When, Value, DecimalField, ExpressionWrapper
from django.db import models
from django.utils import timezone
from django import forms
from django.forms import modelformset_factory
from django.http import JsonResponse, HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None

from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from ledger.models import Party, FinancialRecord, TransactionCategory, Bill, BillTrip, CompanyAccount
from fleet.models import Vehicle, MaintenanceRecord, Tyre
from drivers.models import Driver
from .models import Trip, Route
from .forms import TripForm, RouteForm
import hashlib

@login_required
def reference_data(request):
    """
    Returns a JSON snapshot of all reference data (Parties, Vehicles, Routes, Drivers)
    to be cached in the browser's local storage.
    """
    parties = list(Party.objects.all().order_by('name').values('id', 'name'))
    vehicles = list(Vehicle.objects.filter(status=Vehicle.STATUS_ACTIVE).order_by('registration_plate').values('id', 'registration_plate', 'make_model', 'ownership'))
    routes = list(Route.objects.all().order_by('pickup_location').values('id', 'pickup_location', 'delivery_location', 'route_type', 'default_rate'))
    drivers = list(Driver.objects.select_related('user').all().order_by('user__username').values('id', 'user__username'))
    
    # Generate a robust version key based on the actual content to detect updates
    # We use a stable string representation of the data for hashing
    content_str = f"{parties}{vehicles}{routes}{drivers}"
    version_key = hashlib.md5(content_str.encode()).hexdigest()
    
    return JsonResponse({
        'parties': parties,
        'vehicles': vehicles,
        'routes': routes,
        'drivers': drivers,
        'version': version_key
    })


class BaseTripPermissionMixin:
    """Base mixin for trip permissions"""
    
    def has_manager_permission(self):
        """Check if user has manager dashboard permission"""
        return self.request.user.has_perm('trips.can_view_manager_dashboard')
    
    def has_supervisor_permission(self):
        """Check if user has view all trips permission"""
        return self.request.user.has_perm('trips.can_view_all_trips')
    
    def has_driver_profile(self):
        """Check if user has an associated driver profile"""
        return hasattr(self.request.user, 'driver_profile')

    def has_driver_permission(self):
        """Check if user has driver access (is a driver)"""
        return self.has_driver_profile()
    
    def get_queryset_for_user(self):
        """Filter trips based on user permissions"""
        user = self.request.user
        
        # Admin or user with explicit permission can see all trips
        if user.is_superuser or user.has_perm('trips.can_view_all_trips'):
            return Trip.objects.all()
        
        # Driver can only see their own trips
        if hasattr(user, 'driver_profile'):
            return Trip.objects.filter(driver=user.driver_profile)
        
        # Default: no trips
        return Trip.objects.none()


class TripListView(LoginRequiredMixin, BaseTripPermissionMixin, ListView):
    """
    List view for trips, showing all trips in a list with filters and sorting.
    """
    model = Trip
    template_name = 'trips/trip_list.html'
    context_object_name = 'trips'
    paginate_by = 25
    
    def get_queryset(self):
        """Filter and sort trips based on user input and permissions"""
        queryset = self.get_queryset_for_user().with_payment_info().with_billing_info().select_related(
            'vehicle', 'party', 'driver', 'route'
        ).prefetch_related(
            'bills',
            'bills__category',
            'bills__financial_records',
            'bills__financial_records__category',
            'bills__trips',
            'bills__adjustment_bills',
            'bills__adjustment_bills__category'
        )
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(trip_number__icontains=search) |
                Q(party__name__icontains=search) |
                Q(pickup_location__icontains=search) |
                Q(delivery_location__icontains=search) |
                Q(route__pickup_location__icontains=search) |
                Q(route__delivery_location__icontains=search) |
                Q(vehicle__registration_plate__icontains=search) |
                Q(lr_no__icontains=search)
            ).distinct()

        # LR Number Specific Search
        lr_search = self.request.GET.get('lr_search')
        if lr_search:
            queryset = queryset.filter(lr_no__icontains=lr_search)
        
        # Party filter
        party_id = self.request.GET.get('party')
        if party_id:
            queryset = queryset.filter(party_id=party_id)

        # Status filter (payment-based)
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(annotated_status=status)

        # Billing Status filter
        billing_status = self.request.GET.get('billing_status')
        if billing_status == 'unbilled':
            queryset = queryset.filter(annotated_is_billed=False)
        elif billing_status == 'billed':
            queryset = queryset.filter(annotated_is_billed=True)

        # GST Filter
        gst_filter = self.request.GET.get('gst_filter')
        if gst_filter == 'gst':
            queryset = queryset.filter(annotated_gst_type__in=['GST', 'IGST'])
        elif gst_filter == 'non_gst':
            queryset = queryset.filter(annotated_gst_type='NONE')
            
        # Quick Date filter
        date_filter = self.request.GET.get('date_filter')
        today = timezone.now().date()
        if date_filter == 'today':
            queryset = queryset.filter(date__date=today)
        elif date_filter == 'yesterday':
            queryset = queryset.filter(date__date=today - timedelta(days=1))
        elif date_filter == 'last_7_days':
            queryset = queryset.filter(date__date__gte=today - timedelta(days=7))

        # Date range filtering
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        exact_date = self.request.GET.get('exact_date')
        
        if exact_date:
            try:
                queryset = queryset.filter(date__date=exact_date)
            except (ValueError, TypeError):
                pass
        else:
            if start_date:
                try:
                    queryset = queryset.filter(date__date__gte=start_date)
                except (ValueError, TypeError):
                    pass
            if end_date:
                try:
                    queryset = queryset.filter(date__date__lte=end_date)
                except (ValueError, TypeError):
                    pass

        # Sorting
        sort = self.request.GET.get('sort', '-date')
        sort_mapping = {
            'date': 'date',
            '-date': '-date',
            'trip_number': 'trip_number',
            '-trip_number': '-trip_number',
            'weight': 'weight',
            '-weight': '-weight',
            'revenue': 'annotated_revenue',
            '-revenue': '-annotated_revenue',
        }
        
        if sort in sort_mapping:
            queryset = queryset.order_by(sort_mapping[sort], '-created_at')
        else:
            queryset = queryset.order_by('-date', '-created_at')
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = Trip.PAYMENT_STATUS_CHOICES
        context['current_status'] = self.request.GET.get('status', '')
        context['current_billing_status'] = self.request.GET.get('billing_status', '')
        context['current_gst_filter'] = self.request.GET.get('gst_filter', '')
        context['search_term'] = self.request.GET.get('search', '')
        context['current_lr_search'] = self.request.GET.get('lr_search', '')
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['exact_date'] = self.request.GET.get('exact_date', '')
        context['current_sort'] = self.request.GET.get('sort', '-date')
        context['current_party'] = self.request.GET.get('party', '')
        context['date_filter'] = self.request.GET.get('date_filter', '')
        
        # Summary for the filtered queryset
        queryset = self.get_queryset()
        context['total_weight'] = queryset.aggregate(Sum('weight'))['weight__sum'] or 0
        context['total_count'] = queryset.count()
        
        return context


class TripDetailView(LoginRequiredMixin, BaseTripPermissionMixin, DetailView):
    """
    Detail view for a single trip.
    Uses prefetching + model properties to avoid parser stack overflow.
    """
    model = Trip
    template_name = 'trips/trip_detail.html'
    context_object_name = 'trip'
    
    def get_queryset(self):
        """
        Optimized queryset for detail view.
        We avoid complex annotations (with_payment_info) and use prefetching instead.
        """
        return self.get_queryset_for_user().select_related(
            'vehicle', 'party', 'driver', 'route'
        ).prefetch_related(
            'bills',
            'bills__category',
            'bills__financial_records',
            'bills__financial_records__category',
            'bills__trips',
            'bills__adjustment_bills',
            'bills__adjustment_bills__category',
            'financial_records',
            'financial_records__category',
            'payment_allocations',
            'payment_allocations__financial_record',
            'payment_allocations__financial_record__category'
        )


from django.views.generic import FormView

class TripBulkCreateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    Bulk create view for multiple trips.
    """
    template_name = 'trips/trip_bulk_form.html'
    permission_required = 'trips.add_trip'
    success_url = reverse_lazy('trip-list')
    
    def get_form_class(self):
        return modelformset_factory(
            Trip,
            form=TripForm,
            extra=1,
            can_delete=True
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['queryset'] = Trip.objects.none()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        routes = Route.objects.all().values('id', 'default_rate')
        route_rates = {str(r['id']): float(r['default_rate']) for r in routes}
        context['route_rates'] = route_rates
        return context

    def form_valid(self, form):
        instances = form.save(commit=False)
        for instance in instances:
            instance.created_by = self.request.user
            # Ensure the time portion is set if it's not present (date only)
            if not instance.date:
                 instance.date = timezone.now()
            elif type(instance.date) is datetime.date:
                 current_time = timezone.now().time()
                 instance.date = datetime.combine(instance.date, current_time)
            instance.save()
        messages.success(self.request, f'{len(instances)} Trips created successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "There were errors saving the trips. Please correct the highlighted fields below.")
        return super().form_invalid(form)


class TripCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Create view for new trips.
    """
    model = Trip
    form_class = TripForm
    template_name = 'trips/trip_form.html'
    permission_required = 'trips.add_trip'
    
    def get_initial(self):
        initial = super().get_initial()
        
        # Handle date from GET
        date_str = self.request.GET.get('date')
        if date_str:
            try:
                initial['date'] = datetime.strptime(date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass
                
        # Handle party from GET
        party_id = self.request.GET.get('party')
        if party_id:
            initial['party'] = party_id
            
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        # Set date from GET param if not provided in form (e.g. if field was somehow omitted)
        if not form.cleaned_data.get('date'):
            date_str = self.request.GET.get('date')
            if date_str:
                try:
                    trip_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    current_time = timezone.now().time()
                    form.instance.date = datetime.combine(trip_date, current_time)
                except ValueError:
                    form.instance.date = timezone.now()
            else:
                form.instance.date = timezone.now()

        response = super().form_valid(form)
        messages.success(self.request, 'Trip created successfully!')
        
        # Handle "Save and Add New" options
        if '_save_new_date' in self.request.POST:
            date_str = self.object.date.strftime('%Y-%m-%d')
            return redirect(f"{self.request.path}?date={date_str}")
        elif '_save_new_party' in self.request.POST:
            party_id = self.object.party.id if self.object.party else ''
            return redirect(f"{self.request.path}?party={party_id}")
            
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Provide route default rates for JS auto-population
        routes = Route.objects.all().values('id', 'default_rate')
        route_rates = {str(r['id']): float(r['default_rate']) for r in routes}
        context['route_rates'] = route_rates
        return context

    def get_success_url(self):
        return reverse_lazy('trip-list')


class TripUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Update view for existing trips.
    """
    model = Trip
    form_class = TripForm
    template_name = 'trips/trip_form.html'
    permission_required = 'trips.change_trip'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Trip updated successfully!')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Provide route default rates for JS auto-population
        routes = Route.objects.all().values('id', 'default_rate')
        route_rates = {str(r['id']): float(r['default_rate']) for r in routes}
        context['route_rates'] = route_rates
        return context

    def get_success_url(self):
        return reverse_lazy('trip-detail', kwargs={'pk': self.object.pk})


class TripDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Delete view for trips
    """
    model = Trip
    template_name = 'trips/trip_confirm_delete.html'
    permission_required = 'trips.delete_trip'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Provide route default rates for JS auto-population
        routes = Route.objects.all().values('id', 'default_rate')
        route_rates = {str(r['id']): float(r['default_rate']) for r in routes}
        context['route_rates'] = route_rates
        return context

    def get_success_url(self):
        return reverse_lazy('trip-list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Trip deleted successfully!')
        return super().delete(request, *args, **kwargs)


@login_required
def manager_dashboard(request):
    """
    Manager dashboard - optimized with SQL-level aggregations
    """
    if not (request.user.is_superuser or 
            request.user.has_perm('trips.can_view_manager_dashboard')):
        messages.error(request, 'Access denied. Manager dashboard is only for managers.')
        return redirect('trip-list')
    
    current_date = timezone.now().date()
    current_month = current_date.month
    current_year = current_date.year
    
    # 1. Overall Stats
    trip_stats = Trip.objects.with_payment_info().aggregate(
        total_lifetime=models.Count('id'),
        active_count=models.Count('id', filter=Q(annotated_status__in=['Unpaid', 'Partially Paid'])),
        completed_month=models.Count('id', filter=Q(
            annotated_status='Paid',
            date__month=current_month,
            date__year=current_year
        ))
    )
    
    # 2. Optimized Financials (Aggregated by type in SQL)
    # Filter for the current month
    month_records = FinancialRecord.objects.filter(
        date__month=current_month,
        date__year=current_year
    ).exclude(record_type=FinancialRecord.RECORD_TYPE_INVOICE)

    financial_totals = month_records.aggregate(
        income=Sum('amount', filter=Q(category__type=TransactionCategory.TYPE_INCOME)),
        expenses=Sum('amount', filter=Q(category__type=TransactionCategory.TYPE_EXPENSE) & ~Q(category__name='Deductions'))
    )

    income_this_month = financial_totals['income'] or 0
    expenses_this_month = financial_totals['expenses'] or 0

    # 3. Optimized GST Calculation (Direct SQL aggregation from Bills)
    gst_totals = Bill.objects.filter(
        date__month=current_month,
        date__year=current_year
    ).aggregate(
        total_gst=Sum(
            ExpressionWrapper(
                (F('standard_weight') * F('standard_rate') * F('gst_rate') / Decimal('100')),
                output_field=DecimalField()
            ),
            filter=Q(bill_type='Standard'),
            output_field=DecimalField()
        )
    )
    
    gst_this_month = gst_totals['total_gst'] or 0
    
    # Add trip-based GST (much faster than looping over bill objects)
    trip_gst = Trip.objects.filter(
        bills__date__month=current_month,
        bills__date__year=current_year
    ).aggregate(
        val=Sum(
            ExpressionWrapper(
                Case(
                    When(revenue_type='fixed', then=F('rate_per_ton')),
                    default=F('weight') * F('rate_per_ton'),
                    output_field=DecimalField()
                ) * F('bills__gst_rate') / Decimal('100'),
                output_field=DecimalField()
            )
        )
    )['val'] or 0
    
    gst_this_month += trip_gst

    # 4. Deep Vehicle Utilization & Health
    # We want a list of all vehicles with their performance metrics
    vehicles = Vehicle.objects.exclude(status=Vehicle.STATUS_RETIRED).annotate(
        # Trips
        trips_lifetime=models.Count('trips', distinct=True),
        trips_this_month=models.Count('trips', filter=Q(
            trips__date__month=current_month,
            trips__date__year=current_year
        ), distinct=True),
        
        # Revenue (Subtotal)
        revenue_lifetime=Sum('trips__revenue_cached', distinct=True),
        revenue_this_month=Sum('trips__revenue_cached', filter=Q(
            trips__date__month=current_month,
            trips__date__year=current_year
        ), distinct=True),
        
        # Maintenance Cost (Monthly)
        maintenance_cost_month=Sum('maintenance_records__cost', filter=Q(
            maintenance_records__is_completed=True,
            maintenance_records__completion_date__month=current_month,
            maintenance_records__completion_date__year=current_year
        ), distinct=True),
        
        # Next Maintenance Due
        next_maint_date=models.Min('maintenance_records__expiry_date', filter=Q(
            maintenance_records__is_completed=False
        ))
    ).order_by('registration_plate')

    # 5. Global Alerts
    vehicles_due_maintenance = MaintenanceRecord.objects.filter(
        is_completed=False,
        expiry_date__lte=current_date + timedelta(days=7)
    ).values('vehicle').distinct().count()
    
    recent_trips = Trip.objects.with_payment_info().order_by('-created_at')[:10]
    vehicles_in_maintenance = Vehicle.objects.filter(status=Vehicle.STATUS_MAINTENANCE).count()
    
    context = {
        'today': current_date,
        'total_trips_lifetime': trip_stats['total_lifetime'],
        'active_trips': trip_stats['active_count'],
        'completed_this_month': trip_stats['completed_month'],
        'vehicles_due_maintenance': vehicles_due_maintenance,
        'income_this_month': income_this_month,
        'expenses_this_month': expenses_this_month,
        'net_profit_incl_gst': income_this_month - expenses_this_month,
        'net_profit_excl_gst': (income_this_month - gst_this_month) - expenses_this_month,
        'recent_trips': recent_trips,
        'vehicles_in_maintenance': vehicles_in_maintenance,
        'vehicle_stats': vehicles,
    }
    
    return render(request, 'trips/manager_dashboard.html', context)


@login_required
def get_autocomplete_suggestions(request):
    """
    Returns suggestions for Select2.
    """
    field = request.GET.get('field')
    term = request.GET.get('term', request.GET.get('q', ''))
    
    results = []
    
    if field in ['pickup_location', 'delivery_location']:
        seen_names = set()
        query_filter = {f"{field}__icontains": term} if term else {}
        local_qs = Trip.objects.filter(**query_filter).values_list(field, flat=True).distinct().order_by(field)[:10]
        
        for name in local_qs:
            if name and name not in seen_names:
                results.append({
                    'id': name,
                    'text': f"🕒 {name}",
                    'source': 'history'
                })
                seen_names.add(name)

    elif field == 'tyre_brand':
        from fleet.models import TyreBrand
        qs = TyreBrand.objects.filter(name__icontains=term).order_by('name')[:10]
        results = []
        for brand in qs:
            results.append({
                'id': brand.name, 
                'text': brand.name,
                'price': str(brand.suggestive_price) if brand.suggestive_price else ''
            })
        
        # Also include historical distinct brands if we want, but let's just use TyreBrand since that's the new standard
        if not results:
            qs_old = Tyre.objects.filter(brand__icontains=term).values_list('brand', flat=True).distinct()[:10]
            results = [{'id': x, 'text': x, 'price': ''} for x in qs_old]
    elif field == 'tyre_size':
        qs = Tyre.objects.filter(size__icontains=term).values_list('size', flat=True).distinct()[:10]
        results = [{'id': x, 'text': x} for x in qs]
        
    return JsonResponse({'results': results})

# --- Route Views ---

class RouteListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Route
    template_name = 'trips/route_list.html'
    context_object_name = 'routes'
    permission_required = 'trips.view_route'
    paginate_by = 25

    def get_queryset(self):
        queryset = Route.objects.all().order_by('pickup_location', 'delivery_location')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(pickup_location__icontains=search) |
                Q(delivery_location__icontains=search) |
                Q(route_type__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_term'] = self.request.GET.get('search', '')
        return context

class RouteCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Route
    form_class = RouteForm
    template_name = 'trips/route_form.html'
    permission_required = 'trips.add_route'
    success_url = reverse_lazy('route-list')

    def form_valid(self, form):
        messages.success(self.request, 'Route created successfully.')
        return super().form_valid(form)

class RouteUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Route
    form_class = RouteForm
    template_name = 'trips/route_form.html'
    permission_required = 'trips.change_route'
    success_url = reverse_lazy('route-list')

    def form_valid(self, form):
        messages.success(self.request, 'Route updated successfully.')
        return super().form_valid(form)

class RouteDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Route
    template_name = 'trips/route_confirm_delete.html'
    permission_required = 'trips.delete_route'
    success_url = reverse_lazy('route-list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Route deleted successfully.')
        return super().delete(request, *args, **kwargs)


@login_required
def trip_export_excel(request):
    """
    Generates an Excel sheet for billed trips with vehicle and firm selection.
    """
    if not openpyxl:
        messages.error(request, "Excel export is not available. Please install 'openpyxl'.")
        return redirect('trip-list')

    # Get available vehicles and company accounts (issuers) for selection
    vehicles = Vehicle.objects.all().order_by('registration_plate')
    issuers = CompanyAccount.objects.all().order_by('name')
    
    if request.method == 'POST':
        selected_vehicles = request.POST.getlist('vehicles')
        selected_issuers = request.POST.getlist('issuers')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if not selected_vehicles:
            messages.error(request, "Please select at least one vehicle.")
            return redirect('trip-export-excel')

        # Filter trips: only billed trips for selected vehicles
        trips = Trip.objects.with_billing_info().filter(
            annotated_is_billed=True,
            vehicle_id__in=selected_vehicles
        ).select_related('vehicle', 'party').prefetch_related('bills')

        # Filter by selected issuers (firms) if provided
        if selected_issuers:
            trips = trips.filter(bills__issuer_id__in=selected_issuers).distinct()

        if start_date:
            trips = trips.filter(date__date__gte=start_date)
        if end_date:
            trips = trips.filter(date__date__lte=end_date)

        # Sort by invoice number (numeric part) primarily, then by date
        trips = trips.order_by('bills__bill_no', 'date').distinct()

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Billed Trips Report"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald-500
        center_align = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Header Row
        headers = [
            'Sr no.', 'Invoice No.', 'Date of Invoice', 'Party Name', 'GST No.', 
            'lo-Date', 'Truck no', 'From', 'To', 'Weight', 'Rate', 'Freight', 'GST (18%)', 'Total Taxable Value'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Data Rows
        row_num = 2
        for idx, trip in enumerate(trips, 1):
            # Use smart model properties
            freight = trip.revenue or Decimal('0')
            gst = trip.gst_amount or Decimal('0')
            total_taxable = trip.total_revenue or Decimal('0')

            ws.cell(row=row_num, column=1, value=idx).border = border
            
            # Get associated bill if any
            bill = trip.associated_bill
            ws.cell(row=row_num, column=2, value=bill.bill_number if bill else 'Unbilled').border = border
            ws.cell(row=row_num, column=3, value=bill.date.strftime('%d-%m-%Y') if bill and bill.date else '-').border = border
            ws.cell(row=row_num, column=4, value=trip.party.name if trip.party else 'N/A').border = border
            ws.cell(row=row_num, column=5, value=trip.party.gstin if trip.party else 'N/A').border = border
            ws.cell(row=row_num, column=6, value=trip.date.strftime('%d-%m-%Y')).border = border
            ws.cell(row=row_num, column=7, value=trip.vehicle.registration_plate).border = border
            ws.cell(row=row_num, column=8, value=trip.pickup_location or 'N/A').border = border
            ws.cell(row=row_num, column=9, value=trip.delivery_location or 'N/A').border = border
            ws.cell(row=row_num, column=10, value=float(trip.weight or 0)).border = border
            ws.cell(row=row_num, column=11, value=float(trip.rate_per_ton or 0)).border = border
            ws.cell(row=row_num, column=12, value=float(freight)).border = border
            ws.cell(row=row_num, column=13, value=float(gst)).border = border
            ws.cell(row=row_num, column=14, value=float(total_taxable)).border = border
            
            row_num += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Response
        response = HttpResponse(
            content_type='application/vnd.openpyxl.sheet',
        )
        filename = f"Billed_Trips_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    return render(request, 'trips/trip_export_form.html', {
        'vehicles': vehicles,
        'issuers': issuers
    })
